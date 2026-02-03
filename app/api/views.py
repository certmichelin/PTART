import datetime
from django.conf import settings

from django.contrib.auth import authenticate
from django.contrib.auth.password_validation import validate_password

from django.core.exceptions import ValidationError

from django.http import HttpResponse, JsonResponse

from django.views.decorators.csrf import csrf_exempt

from rest_framework import status
from rest_framework.decorators import action, renderer_classes
from rest_framework.decorators import api_view
from rest_framework.renderers import BaseRenderer, TemplateHTMLRenderer
from rest_framework.response import Response
from rest_framework.authtoken.models import Token

from openpyxl import Workbook, styles
from openpyxl.styles import Alignment
from openpyxl.writer.excel import save_virtual_workbook

import cairosvg
import jinja2
import os
import pypandoc
import random
import re
import requests
import zipfile

from ptart.models import (
    Flag,
    Hit,
    Assessment,
    Project,
    Template,
    Comment,
    HitReference,
    Host,
    Service,
    Screenshot,
    Attachment,
    Cvss3,
    Cvss4,
    Case,
    Module,
    Methodology,
    Label,
    CWEs,
    CWE,
    AttackScenario,
    Recommendation,
    Tool,
    RetestCampaign,
    RetestHit,
    RetestScreenshot,
    User,
)

from api.decorators import ptart_authentication

from .serializers import (
    FlagSerializer,
    HitSerializer,
    AssessmentSerializer,
    ProjectSerializer,
    TemplateSerializer,
    HostSerializer,
    ServiceSerializer,
    ScreenshotSerializer,
    AttachmentSerializer,
    CommentSerializer,
    HitReferenceSerializer,
    Cvss3Serializer,
    Cvss4Serializer,
    CaseSerializer,
    ModuleSerializer,
    MethodologySerializer,
    LabelSerializer,
    CWEsSerializer,
    CWESerializer,
    AttackScenarioSerializer,
    RecommendationSerializer,
    ToolSerializer,
    RetestCampaignSerializer,
    RetestHitSerializer,
    RetestScreenshotSerializer,
    UserSerializer
)

#
# Image renderer for Screenshots.
#
class ImageRenderer(BaseRenderer):
    def render(self, data, media_type="image/png", renderer_context=None):
        return data

#
# JSON renderer for Burp config file.
#
class JsonRenderer(BaseRenderer):
    def render(self, data, media_type="application/json", renderer_context=None):
        return data


#
# Binary renderer for Attachments.
#
class BinaryRenderer(BaseRenderer):
    def render(
        self, data, media_type="application/octet-stream", renderer_context=None
    ):
        return data

#
# Xlsx renderer for Excel reports.
#
class XlsxRenderer(BaseRenderer):
    media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    format_suffix = "xlsx"

    def render(
        self, data, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", renderer_context=None
    ):
        return data

#
# Zip renderer for LaTeX reports.
#
class ZipRenderer(BaseRenderer):
    media_type = "application/zip"
    format_suffix = "zip"

    def render(
        self, data, media_type="application/zip", renderer_context=None
    ):
        return data

@csrf_exempt
@ptart_authentication
@api_view(["GET", "PATCH", "PUT", "DELETE"])
def flag(request, pk):
    return item(request, pk, Flag, FlagSerializer)


@csrf_exempt
@ptart_authentication
@api_view(["GET", "POST"])
def flags(request):
    return items(request, Flag, FlagSerializer)


@csrf_exempt
@ptart_authentication
@api_view(["GET", "PATCH", "PUT", "DELETE"])
def hit(request, pk):
    return item(request, pk, Hit, HitSerializer)


@csrf_exempt
@ptart_authentication
@api_view(["GET", "POST"])
def hits(request):
    return items(request, Hit, HitSerializer)


@csrf_exempt
@ptart_authentication
@api_view(["GET", "PATCH", "PUT", "DELETE"])
def attackscenario(request, pk):
    return item(request, pk, AttackScenario, AttackScenarioSerializer)


@csrf_exempt
@ptart_authentication
@api_view(["GET", "POST"])
def attackscenarios(request):
    return items(request, AttackScenario, AttackScenarioSerializer)


@csrf_exempt
@ptart_authentication
@api_view(["GET", "PATCH", "PUT", "DELETE"])
def recommendation(request, pk):
    return item(request, pk, Recommendation, RecommendationSerializer)


@csrf_exempt
@ptart_authentication
@api_view(["GET", "POST"])
def recommendations(request):
    return items(request, Recommendation, RecommendationSerializer)


@csrf_exempt
@ptart_authentication
@api_view(["GET", "PATCH", "PUT", "DELETE"])
def label(request, pk):
    return item(request, pk, Label, LabelSerializer)


@csrf_exempt
@ptart_authentication
@api_view(["GET", "POST"])
def labels(request):
    return items(request, Label, LabelSerializer)


@csrf_exempt
@ptart_authentication
@api_view(["GET", "PATCH", "PUT", "DELETE"])
def tool(request, pk):
    return item(request, pk, Tool, ToolSerializer)


@csrf_exempt
@ptart_authentication
@api_view(["GET", "POST"])
def tools(request):
    return items(request, Tool, ToolSerializer)


@csrf_exempt
@ptart_authentication
@api_view(["GET", "DELETE"])
def comment(request, pk):
    return item(request, pk, Comment, CommentSerializer)


@csrf_exempt
@ptart_authentication
@api_view(["GET", "DELETE"])
def hit_reference(request, pk):
    return item(request, pk, HitReference, HitReferenceSerializer)


@csrf_exempt
@ptart_authentication
@api_view(["DELETE"])
def screenshot(request, pk):
    return item(request, pk, Screenshot, ScreenshotSerializer)


@csrf_exempt
@ptart_authentication
@api_view(["GET", "POST"])
def screenshots(request):
    return items(request, Screenshot, ScreenshotSerializer)


@csrf_exempt
@ptart_authentication
@api_view(["DELETE"])
def attachment(request, pk):
    return item(request, pk, Attachment, AttachmentSerializer)


@csrf_exempt
@ptart_authentication
@api_view(["GET", "POST"])
def attachments(request):
    return items(request, Attachment, AttachmentSerializer)


@csrf_exempt
@ptart_authentication
@api_view(["GET", "PATCH", "PUT", "DELETE"])
def assessment(request, pk):
    return item(request, pk, Assessment, AssessmentSerializer)


@csrf_exempt
@ptart_authentication
@api_view(["GET", "POST"])
def assessments(request):
    return items(request, Assessment, AssessmentSerializer)


@csrf_exempt
@ptart_authentication
@api_view(["GET", "PATCH", "PUT", "DELETE"])
def host(request, pk):
    return item(request, pk, Host, HostSerializer)


@csrf_exempt
@ptart_authentication
@api_view(["GET", "POST"])
def hosts(request):
    return items(request, Host, HostSerializer)


@csrf_exempt
@ptart_authentication
@api_view(["GET", "PATCH", "PUT", "DELETE"])
def service(request, pk):
    return item(request, pk, Service, ServiceSerializer)


@csrf_exempt
@ptart_authentication
@api_view(["GET", "POST"])
def services(request):
    return items(request, Service, ServiceSerializer)


@csrf_exempt
@ptart_authentication
@api_view(["GET", "PATCH", "PUT", "DELETE"])
def project(request, pk):
    return item(request, pk, Project, ProjectSerializer)


@csrf_exempt
@ptart_authentication
@api_view(["GET", "POST"])
def projects(request):
    return items(request, Project, ProjectSerializer)


@csrf_exempt
@ptart_authentication
@api_view(["GET", "PATCH", "PUT", "DELETE"])
def template(request, pk):
    return item(request, pk, Template, TemplateSerializer)


@csrf_exempt
@ptart_authentication
@api_view(["GET", "POST"])
def templates(request):
    return items(request, Template, TemplateSerializer)


@csrf_exempt
@ptart_authentication
@api_view(["GET", "PATCH", "PUT", "DELETE"])
def case(request, pk):
    return item(request, pk, Case, CaseSerializer)


@csrf_exempt
@ptart_authentication
@api_view(["GET", "POST"])
def cases(request):
    return items(request, Case, CaseSerializer)


@csrf_exempt
@ptart_authentication
@api_view(["GET", "PATCH", "PUT", "DELETE"])
def module(request, pk):
    return item(request, pk, Module, ModuleSerializer)


@csrf_exempt
@ptart_authentication
@api_view(["GET", "POST"])
def modules(request):
    return items(request, Module, ModuleSerializer)


@csrf_exempt
@ptart_authentication
@api_view(["GET", "PATCH", "PUT", "DELETE"])
def methodology(request, pk):
    return item(request, pk, Methodology, MethodologySerializer)


@csrf_exempt
@ptart_authentication
@api_view(["GET", "POST"])
def methodologies(request):
    return items(request, Methodology, MethodologySerializer)


@csrf_exempt
@ptart_authentication
@api_view(["GET", "PATCH", "PUT"])
def cwe_list(request, pk):
    return item(request, pk, CWEs, CWEsSerializer)


@csrf_exempt
@ptart_authentication
@api_view(["GET"])
def cwe_lists(request):
    return items(request, CWEs, CWEsSerializer)


@csrf_exempt
@ptart_authentication
@api_view(["GET"])
def cwe(request, pk):
    return item(request, pk, CWE, CWESerializer)


@csrf_exempt
@ptart_authentication
@api_view(["GET"])
def cwes(request):
    return items(request, CWE, CWESerializer)


@csrf_exempt
@ptart_authentication
@api_view(["GET", "PATCH", "PUT", "DELETE"])
def retestcampaign(request, pk):
    return item(request, pk, RetestCampaign, RetestCampaignSerializer)


@csrf_exempt
@ptart_authentication
@api_view(["GET", "POST"])
def retestcampaigns(request):
    return items(request, RetestCampaign, RetestCampaignSerializer)


@csrf_exempt
@ptart_authentication
@api_view(["GET", "PATCH", "PUT", "DELETE"])
def retesthit(request, pk):
    return item(request, pk, RetestHit, RetestHitSerializer)


@csrf_exempt
@ptart_authentication
@api_view(["GET", "POST"])
def retesthits(request):
    return items(request, RetestHit, RetestHitSerializer)


@csrf_exempt
@ptart_authentication
@api_view(["DELETE"])
def retestscreenshot(request, pk):
    return item(request, pk, RetestScreenshot, RetestScreenshotSerializer)


@csrf_exempt
@ptart_authentication
@api_view(["GET", "POST"])
def retestscreenshots(request):
    return items(request, RetestScreenshot, RetestScreenshotSerializer)

@csrf_exempt
@ptart_authentication
@api_view(["GET"])
def users(request):
    response = None
    userList = User.objects.filter(is_active=True)
    if request.method == "GET":
        response = Response(
            UserSerializer(userList, many=True).data
        )
    return response

@csrf_exempt
@ptart_authentication
@api_view(["POST"])
def load_module(request, pk, assessmentId):
    response = None
    try:
        assessment = Assessment.objects.get(pk=assessmentId)
        if assessment.is_user_can_edit(request.user):
            module = Module.objects.get(pk=pk)
            cases = Case.objects.filter(module=module)
            flags = []
            for case in cases:
                note = "Module: " + module.name + "\n\n" + case.description
                flag = Flag(
                    title=case.name,
                    note=note,
                    assessment=assessment,
                    assignee=request.user,
                )
                flag.save()
                flags.append(flag)
            response = Response(
                FlagSerializer(flags, many=True).data, status=status.HTTP_201_CREATED
            )
        else:
            response = Response(status=status.HTTP_403_FORBIDDEN)
    except Module.DoesNotExist:
        response = Response(status=status.HTTP_404_NOT_FOUND)
    except Assessment.DoesNotExist:
        response = Response(status=status.HTTP_404_NOT_FOUND)
    return response


@csrf_exempt
@ptart_authentication
@api_view(["PUT"])
def screenshot_order(request, pk, order):
    response = None
    try:
        item = Screenshot.objects.get(pk=pk)
        if item.is_user_can_edit(request.user):
            order = int(order)
            if (
                item.order == order
                or order < 0
                or order >= item.hit.screenshot_set.count()
            ):
                response = Response(status=status.HTTP_400_BAD_REQUEST)
            else:
                # Reorder all the screenshot of the hit.
                if item.order < order:
                    # Move the screenshot down.
                    for screenshot in item.hit.screenshot_set.all():
                        if screenshot.order > item.order and screenshot.order <= order:
                            screenshot.order = screenshot.order - 1
                            screenshot.save(update_fields=["order"])
                else:
                    # Move the screenshot up.
                    for screenshot in item.hit.screenshot_set.all():
                        if screenshot.order < item.order and screenshot.order >= order:
                            screenshot.order = screenshot.order + 1
                            screenshot.save(update_fields=["order"])

                # Save the new order of the screenshot.
                item.order = order
                item.save(update_fields=["order"])
                response = Response(
                    ScreenshotSerializer(item).data, status=status.HTTP_202_ACCEPTED
                )
        else:
            response = Response(status=status.HTTP_403_FORBIDDEN)
    except Screenshot.DoesNotExist:
        response = Response(status=status.HTTP_404_NOT_FOUND)
    return response


@csrf_exempt
@api_view(["GET"])
@ptart_authentication
def screenshot_raw(request, pk):
    try:
        item = Screenshot.objects.get(pk=pk)
        if item.is_user_can_view(request.user):
            raw_data = item.get_raw_data()
            if raw_data is not None:
                return HttpResponse(raw_data, content_type="image/png")
            else:
                return HttpResponse(status=status.HTTP_404_NOT_FOUND)
        else:
            return HttpResponse(status=status.HTTP_403_FORBIDDEN)
    except Screenshot.DoesNotExist:
        return HttpResponse(status=status.HTTP_404_NOT_FOUND)


@csrf_exempt
@ptart_authentication
@api_view(["PUT"])
def retestscreenshot_order(request, pk, order):
    response = None
    try:
        item = RetestScreenshot.objects.get(pk=pk)
        if item.is_user_can_edit(request.user):
            order = int(order)
            if (
                item.order == order
                or order < 0
                or order >= item.retest_hit.retestscreenshot_set.count()
            ):
                response = Response(status=status.HTTP_400_BAD_REQUEST)
            else:
                # Reorder all the screenshot of the retest hit.
                if item.order < order:
                    # Move the screenshot down.
                    for screenshot in item.retest_hit.retestscreenshot_set.all():
                        if screenshot.order > item.order and screenshot.order <= order:
                            screenshot.order = screenshot.order - 1
                            screenshot.save(update_fields=["order"])
                else:
                    # Move the screenshot up.
                    for screenshot in item.retest_hit.retestscreenshot_set.all():
                        if screenshot.order < item.order and screenshot.order >= order:
                            screenshot.order = screenshot.order + 1
                            screenshot.save(update_fields=["order"])

                # Save the new order of the screenshot.
                item.order = order
                item.save(update_fields=["order"])
                response = Response(
                    RetestScreenshotSerializer(item).data,
                    status=status.HTTP_202_ACCEPTED,
                )
        else:
            response = Response(status=status.HTTP_403_FORBIDDEN)
    except RetestScreenshot.DoesNotExist:
        response = Response(status=status.HTTP_404_NOT_FOUND)
    return response


@csrf_exempt
@api_view(["GET"])
@ptart_authentication
def retestscreenshot_raw(request, pk):
    try:
        item = RetestScreenshot.objects.get(pk=pk)
        if item.is_user_can_view(request.user):
            raw_data = item.get_raw_data()
            if raw_data is not None:
                return HttpResponse(raw_data, content_type="image/png")
            else:
                return HttpResponse(status=status.HTTP_404_NOT_FOUND)
        else:
            return HttpResponse(status=status.HTTP_403_FORBIDDEN)
    except RetestScreenshot.DoesNotExist:
        return HttpResponse(status=status.HTTP_404_NOT_FOUND)


@csrf_exempt
@api_view(["GET"])
@ptart_authentication
def attachment_raw(request, pk):
    try:
        item = Attachment.objects.get(pk=pk)
        if item.is_user_can_view(request.user):
            response = HttpResponse(item.get_raw_data(), content_type="application/octet-stream")
            response["Content-Disposition"] = (
                "attachment; filename=" + item.attachment_name
            )
            return response
        else:
            return HttpResponse(status=status.HTTP_403_FORBIDDEN)
    except Attachment.DoesNotExist:
        return HttpResponse(status=status.HTTP_404_NOT_FOUND)


@csrf_exempt
@ptart_authentication
@api_view(["POST"])
def cvss3(request):
    serializer = Cvss3Serializer(data=request.data)
    if serializer.is_valid():
        cvss = Cvss3(**serializer.validated_data)
        cvss.compute_cvss_value()
        response = Response(Cvss3Serializer(cvss).data, status=status.HTTP_201_CREATED)
    else:
        response = Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    return response


@csrf_exempt
@ptart_authentication
@api_view(["POST"])
def cvss4(request):
    serializer = Cvss4Serializer(data=request.data)
    if serializer.is_valid():
        cvss = Cvss4(**serializer.validated_data)
        cvss.compute_cvss_value()
        response = Response(Cvss4Serializer(cvss).data, status=status.HTTP_201_CREATED)
    else:
        response = Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    return response


@csrf_exempt
@ptart_authentication
@api_view(["POST", "DELETE"])
def cvss3_hit(request, pk):
    response = None
    try:
        hit = Hit.objects.get(pk=pk)
        if hit.is_user_can_edit(request.user):

            if request.method == "DELETE":
                # Delete the cvss attached to the hit.
                if hit.cvss3 is not None:
                    Cvss3.objects.get(pk=hit.cvss3.id).delete()
                response = Response(status=status.HTTP_200_OK)

            else:
                serializer = Cvss3Serializer(data=request.data)
                if serializer.is_valid():
                    # Create the new CVSS.
                    serializer.save()
                    cvss = Cvss3.objects.get(pk=serializer.data["id"])
                    cvss.compute_cvss_value()
                    cvss.save(update_fields=["decimal_value"])

                    # This condition prevent memory leak in DB.
                    if hit.cvss3 is not None:
                        Cvss3.objects.get(pk=hit.cvss3.id).delete()

                    hit.cvss3 = cvss
                    hit.save(update_fields=["cvss3"])
                    response = Response(status=status.HTTP_201_CREATED)
                else:
                    response = Response(
                        serializer.errors, status=status.HTTP_400_BAD_REQUEST
                    )
        else:
            response = Response(status=status.HTTP_403_FORBIDDEN)
    except Hit.DoesNotExist:
        response = Response(status=status.HTTP_404_NOT_FOUND)
    return response


@csrf_exempt
@ptart_authentication
@api_view(["POST", "DELETE"])
def cvss4_hit(request, pk):
    response = None
    try:
        hit = Hit.objects.get(pk=pk)
        if hit.is_user_can_edit(request.user):

            if request.method == "DELETE":
                # Delete the cvss attached to the hit.
                if hit.cvss4 is not None:
                    Cvss4.objects.get(pk=hit.cvss4.id).delete()
                response = Response(status=status.HTTP_200_OK)

            else:
                serializer = Cvss4Serializer(data=request.data)
                if serializer.is_valid():
                    # Create the new CVSS.
                    serializer.save()
                    cvss = Cvss4.objects.get(pk=serializer.data["id"])
                    cvss.compute_cvss_value()
                    cvss.save(update_fields=["decimal_value"])

                    # This condition prevent memory leak in DB.
                    if hit.cvss4 is not None:
                        Cvss4.objects.get(pk=hit.cvss4.id).delete()

                    hit.cvss4 = cvss
                    hit.save(update_fields=["cvss4"])
                    response = Response(status=status.HTTP_201_CREATED)
                else:
                    response = Response(
                        serializer.errors, status=status.HTTP_400_BAD_REQUEST
                    )
        else:
            response = Response(status=status.HTTP_403_FORBIDDEN)
    except Hit.DoesNotExist:
        response = Response(status=status.HTTP_404_NOT_FOUND)
    return response


@csrf_exempt
@ptart_authentication
@api_view(["POST", "GET"])
def comments(request, pk):
    response = None
    try:
        hit = Hit.objects.get(pk=pk)
        if request.method == "GET":
            if hit.is_user_can_view(request.user):
                response = Response(
                    CommentSerializer(hit.comment_set.all(), many=True).data
                )
            else:
                response = Response(status=status.HTTP_403_FORBIDDEN)
        else:
            text = request.data["text"]
            if text is not None and text.strip():
                try:
                    comment = Comment(text=text, hit=hit, author=request.user)
                    if comment.is_user_can_create(request.user):
                        comment.save()
                        response = Response(
                            CommentSerializer(comment).data, status=status.HTTP_200_OK
                        )
                    else:
                        response = Response(status=status.HTTP_403_FORBIDDEN)
                except Hit.DoesNotExist:
                    response = Response(status=status.HTTP_404_NOT_FOUND)
            else:
                response = Response(status=status.HTTP_400_BAD_REQUEST)
    except Hit.DoesNotExist:
        response = Response(status=status.HTTP_404_NOT_FOUND)
    return response


@csrf_exempt
@ptart_authentication
@api_view(["POST", "GET"])
def hit_references(request, pk):
    response = None
    try:
        hit = Hit.objects.get(pk=pk)
        if request.method == "GET":
            if hit.is_user_can_view(request.user):
                response = Response(
                    HitReferenceSerializer(hit.hitreference_set.all(), many=True).data
                )
            else:
                response = Response(status=status.HTTP_403_FORBIDDEN)
        else:
            name = request.data["name"]
            url = request.data["url"]
            if (
                name is not None
                and name.strip()
                and url is not None
                and url.strip()
                and url.startswith("http")
            ):
                try:
                    hitreference = HitReference(name=name, url=url, hit=hit)
                    if hitreference.is_user_can_create(request.user):
                        hitreference.save()
                        response = Response(
                            HitReferenceSerializer(hitreference).data,
                            status=status.HTTP_200_OK,
                        )
                    else:
                        response = Response(status=status.HTTP_403_FORBIDDEN)
                except Hit.DoesNotExist:
                    response = Response(status=status.HTTP_404_NOT_FOUND)
            else:
                response = Response(status=status.HTTP_400_BAD_REQUEST)
    except Hit.DoesNotExist:
        response = Response(status=status.HTTP_404_NOT_FOUND)
    return response


@csrf_exempt
@ptart_authentication
@api_view(["POST"])
def markFlagAsDone(request, pk):
    response = None
    try:
        flag = Flag.objects.get(pk=pk)
        if flag.is_user_can_edit(request.user):
            flag.done = True
            flag.save()
            response = Response(FlagSerializer(flag).data, status=status.HTTP_200_OK)
        else:
            response = Response(status=status.HTTP_403_FORBIDDEN)
    except Flag.DoesNotExist:
        response = Response(status=status.HTTP_404_NOT_FOUND)
    return response


@csrf_exempt
@ptart_authentication
@api_view(["GET"])
def project_burp_configuration(request, pk):
    response = None
    try:
        project = Project.objects.get(pk=pk)
        if project.is_user_can_view(request.user):
            configuration = {"target": {"scope": {"advanced_mode": "true"}}}
            targets = []
            for target in project.scope.splitlines():
                burp_target = {}
                burp_target["enabled"] = True

                # Manage protocol
                protocol = "any"
                if target.startswith("https://"):
                    protocol = "https"
                    target = target.replace("https://", "")
                elif target.startswith("http://"):
                    protocol = "http"
                    target = target.replace("http://", "")
                burp_target["protocol"] = protocol

                # Manage port
                if ":" in target:
                    port = target.split(":")[1]
                    if "/" in port:
                        port = port.split("/")[0]
                    burp_target["port"] = "^" + port + "$"
                    target = target.replace(":" + port, "")

                # Manage file
                if "/" in target:
                    file = target.partition("/")[2]
                    burp_target["file"] = "^" + file + "$"
                    target = target.replace("/" + file, "")

                # Manage Target
                target = target.replace(".", "\\.")
                target = target.replace("*", ".*")
                burp_target["host"] = "^" + target + "$"

                targets.append(burp_target)
            configuration["target"]["scope"]["include"] = targets
            response = Response(configuration)
            response.accepted_renderer = JsonRenderer()
            response.accepted_media_type = "application/json"
            response.renderer_context = {}
            response["Content-Disposition"] = (
                "attachment; filename=" + project.name + "_burp_configuration.json"
            )
        else:
            response = Response(status=status.HTTP_403_FORBIDDEN)
    except Flag.DoesNotExist:
        response = Response(status=status.HTTP_404_NOT_FOUND)
    return response


@csrf_exempt
@ptart_authentication
@api_view(http_method_names=["GET"])
@renderer_classes([XlsxRenderer])
def project_xlsx(request, pk):
    response = None
    try:
        project = Project.objects.get(pk=pk)
        if project.is_user_can_view(request.user):

            # --------------------------------------------------------------------------
            # Main sheet.
            # --------------------------------------------------------------------------

            wb = Workbook()
            ws = wb.active
            ws.title = "Synthesis"

            # Define column size
            wb.active.column_dimensions["A"].width = 28
            wb.active.column_dimensions["B"].width = 10
            wb.active.column_dimensions["C"].width = 10
            wb.active.column_dimensions["D"].width = 17
            wb.active.column_dimensions["E"].width = 50
            wb.active.column_dimensions["F"].width = 40
            wb.active.column_dimensions["G"].width = 12
            wb.active.column_dimensions["H"].width = 50
            wb.active.column_dimensions["I"].width = 70

            # Add project data.
            ws["A1"] = "Project Name:"
            ws["A2"] = "Client:"
            ws["A3"] = "Date:"
            ws["A4"] = "Auditors:"
            ws["A5"] = "Reviewers:"
            ws["B1"] = project.name
            ws["B2"] = project.client
            if project.start_date is not None and project.end_date is not None:
                ws["B3"] = (
                    "From " + str(project.start_date) + " To " + str(project.end_date)
                )
            else:
                ws["B3"] = project.added

            # Construct the auditor string
            pentester_str = ""
            previous = ""
            for pentester in project.pentesters.all():
                pentester_str = "{}{}{} - {} {}".format(
                    pentester_str,
                    previous,
                    pentester.username,
                    pentester.first_name,
                    pentester.last_name,
                )
                previous = ", "
            ws["B4"] = pentester_str

            # Construct the reviewer string
            reviewer_str = ""
            previous = ""
            for reviewer in project.reviewers.all():
                reviewer_str = "{}{}{} - {} {}".format(
                    reviewer_str,
                    previous,
                    reviewer.username,
                    reviewer.first_name,
                    reviewer.last_name,
                )
                previous = ", "
            ws["B5"] = reviewer_str

            # Beautify project data
            ws.merge_cells("B1:I1")
            ws.merge_cells("B2:I2")
            ws.merge_cells("B3:I3")
            ws.merge_cells("B4:I4")
            ws.merge_cells("B5:I5")

            projectHeaderStyle = styles.NamedStyle(name="project_header_style")
            projectHeaderStyle.font = styles.Font(
                name="Calibri", size=14, bold=True, color="000000"
            )
            projectHeaderStyle.fill = styles.PatternFill(
                patternType="solid", fgColor="00B0F0"
            )
            projectHeaderStyle.alignment = styles.Alignment(horizontal="left")
            ws["A1"].style = projectHeaderStyle
            ws["A2"].style = projectHeaderStyle
            ws["A3"].style = projectHeaderStyle
            ws["A4"].style = projectHeaderStyle
            ws["A5"].style = projectHeaderStyle

            projectValueStyle = styles.NamedStyle(name="project_value_style")
            projectValueStyle.font = styles.Font(
                name="Calibri", size=14, italic=True, color="000000"
            )
            projectValueStyle.fill = styles.PatternFill(
                patternType="solid", fgColor="00B0F0"
            )
            projectValueStyle.alignment = styles.Alignment(horizontal="left")

            ws["B1"].style = projectValueStyle
            ws["B2"].style = projectValueStyle
            ws["B3"].style = projectValueStyle
            ws["B4"].style = projectValueStyle
            ws["B5"].style = projectValueStyle
            ws["B3"].number_format = "YYYY MMM DD"

            # Add column header.
            ws["A7"] = "Assessment"
            ws["B7"] = "Sev"
            ws["C7"] = "CVSSv3" if project.cvss_type == 3 else "CVSSv4"
            ws["D7"] = "ID"
            ws["E7"] = "Title"
            ws["F7"] = "Asset"
            ws["G7"] = "Fix Compl."
            ws["H7"] = "Labels"
            ws["I7"] = "CWE(s)"

            columnHeaderStyle = styles.NamedStyle(name="column_header_style")
            columnHeaderStyle.font = styles.Font(
                name="Calibri", size=12, bold=True, color="000000"
            )
            columnHeaderStyle.fill = styles.PatternFill(
                patternType="solid", fgColor="92D050"
            )
            columnHeaderStyle.alignment = styles.Alignment(horizontal="center")

            ws["A7"].style = columnHeaderStyle
            ws["B7"].style = columnHeaderStyle
            ws["C7"].style = columnHeaderStyle
            ws["D7"].style = columnHeaderStyle
            ws["E7"].style = columnHeaderStyle
            ws["F7"].style = columnHeaderStyle
            ws["G7"].style = columnHeaderStyle
            ws["H7"].style = columnHeaderStyle
            ws["I7"].style = columnHeaderStyle

            # Fill the report
            criticalStyle = styles.NamedStyle(name="critical_style")
            criticalStyle.font = styles.Font(name="OCR A Extended", color="FFFFFF")
            criticalStyle.fill = styles.PatternFill(
                patternType="solid", fgColor="343a40"
            )

            highStyle = styles.NamedStyle(name="high_style")
            highStyle.font = styles.Font(name="OCR A Extended", color="FFFFFF")
            highStyle.fill = styles.PatternFill(patternType="solid", fgColor="dc3545")

            mediumStyle = styles.NamedStyle(name="medium_style")
            mediumStyle.font = styles.Font(name="OCR A Extended", color="212529")
            mediumStyle.fill = styles.PatternFill(patternType="solid", fgColor="ffc107")

            lowStyle = styles.NamedStyle(name="low_style")
            lowStyle.font = styles.Font(name="OCR A Extended", color="FFFFFF")
            lowStyle.fill = styles.PatternFill(patternType="solid", fgColor="28a745")

            infoStyle = styles.NamedStyle(name="info_style")
            infoStyle.font = styles.Font(name="OCR A Extended", color="FFFFFF")
            infoStyle.fill = styles.PatternFill(patternType="solid", fgColor="6c757d")

            line = 8
            for assessment in project.assessment_set.all():
                for hit in assessment.displayable_hits():
                    ws.cell(row=line, column=1).value = assessment.name
                    ws.cell(row=line, column=2).value = "P{}".format(hit.severity)
                    ws.cell(row=line, column=3).value = hit.get_cvss_value()
                    ws.cell(row=line, column=4).value = hit.get_unique_id()
                    ws.cell(row=line, column=5).value = hit.title
                    ws.cell(row=line, column=6).value = hit.asset
                    ws.cell(row=line, column=7).value = hit.get_fix_complexity_str()

                    label_str = ""
                    previous = ""
                    for label in hit.labels.all():
                        label_str = "{}{}{}".format(label_str, previous, label.title)
                        previous = "\n"
                    ws.cell(row=line, column=8).value = label_str

                    cwe_str = ""
                    previous = ""
                    for cwe in hit.cwes.all():
                        cwe_str = "{}{}{}".format(cwe_str, previous, cwe)
                        previous = "\n"
                    ws.cell(row=line, column=9).value = cwe_str

                    # Apply style from value.
                    if hit.severity == 1:
                        ws.cell(row=line, column=2).style = criticalStyle
                    elif hit.severity == 2:
                        ws.cell(row=line, column=2).style = highStyle
                    elif hit.severity == 3:
                        ws.cell(row=line, column=2).style = mediumStyle
                    elif hit.severity == 4:
                        ws.cell(row=line, column=2).style = lowStyle
                    elif hit.severity == 5:
                        ws.cell(row=line, column=2).style = infoStyle

                    if hit.fix_complexity == 1:
                        ws.cell(row=line, column=7).style = highStyle
                    elif hit.fix_complexity == 2:
                        ws.cell(row=line, column=7).style = mediumStyle
                    elif hit.fix_complexity == 3:
                        ws.cell(row=line, column=7).style = lowStyle
                    else:
                        ws.cell(row=line, column=7).style = infoStyle

                    # For the moment, CVSS3 and CVSS4 share the same classification values.
                    try:
                        if float(hit.get_cvss_value()) < 0.1:
                            ws.cell(row=line, column=3).style = infoStyle
                        elif float(hit.get_cvss_value()) < 4.0:
                            ws.cell(row=line, column=3).style = lowStyle
                        elif float(hit.get_cvss_value()) < 7.0:
                            ws.cell(row=line, column=3).style = mediumStyle
                        elif float(hit.get_cvss_value()) < 9.0:
                            ws.cell(row=line, column=3).style = highStyle
                        else:
                            ws.cell(row=line, column=3).style = criticalStyle
                    except ValueError:
                        ws.cell(row=line, column=3).style = infoStyle

                    ws.cell(row=line, column=1).alignment = Alignment(
                        vertical="center", horizontal="left"
                    )
                    ws.cell(row=line, column=2).alignment = Alignment(
                        vertical="center", horizontal="center"
                    )
                    ws.cell(row=line, column=3).alignment = Alignment(
                        vertical="center", horizontal="center"
                    )
                    ws.cell(row=line, column=4).alignment = Alignment(
                        vertical="center", horizontal="center"
                    )
                    ws.cell(row=line, column=5).alignment = Alignment(
                        vertical="top", horizontal="left", wrap_text=True
                    )
                    ws.cell(row=line, column=6).alignment = Alignment(
                        vertical="top", horizontal="left", wrap_text=True
                    )
                    ws.cell(row=line, column=7).alignment = Alignment(
                        vertical="center", horizontal="center"
                    )
                    ws.cell(row=line, column=8).alignment = Alignment(
                        vertical="top", wrap_text=True
                    )
                    ws.cell(row=line, column=9).alignment = Alignment(
                        vertical="top", wrap_text=True
                    )

                    line = line + 1

            ws.auto_filter.ref = "A7:I{}".format(line)

            # --------------------------------------------------------------------------
            # Recommendations
            # --------------------------------------------------------------------------

            if project.recommendation_set.all():
                recommendations_ws = wb.create_sheet()
                recommendations_ws.title = "Recommendations"

                # Define column size
                recommendations_ws.column_dimensions["A"].width = 50
                recommendations_ws.column_dimensions["B"].width = 120

                recommendations_ws["A1"].style = columnHeaderStyle
                recommendations_ws["B1"].style = columnHeaderStyle

                # Add column header.
                recommendations_ws["A1"] = "Name"
                recommendations_ws["B1"] = "Body"

                line = 2
                for recommendation in project.recommendation_set.all():
                    recommendations_ws.cell(row=line, column=1).value = (
                        recommendation.name
                    )
                    recommendations_ws.cell(row=line, column=2).value = (
                        recommendation.body
                    )
                    recommendations_ws.cell(row=line, column=1).alignment = Alignment(
                        horizontal="left", vertical="top"
                    )
                    recommendations_ws.cell(row=line, column=2).alignment = Alignment(
                        wrap_text=True
                    )
                    line = line + 1

            # --------------------------------------------------------------------------
            # Restest Campaigns
            # --------------------------------------------------------------------------
            if project.retestcampaign_set.all():
                retestcampaigns_ws = wb.create_sheet()
                retestcampaigns_ws.title = "Retest Campaigns"

                # Define column size
                retestcampaigns_ws.column_dimensions["A"].width = 40
                retestcampaigns_ws.column_dimensions["B"].width = 15
                retestcampaigns_ws.column_dimensions["C"].width = 15
                retestcampaigns_ws.column_dimensions["D"].width = 7
                retestcampaigns_ws.column_dimensions["E"].width = 7
                retestcampaigns_ws.column_dimensions["F"].width = 7
                retestcampaigns_ws.column_dimensions["G"].width = 7
                retestcampaigns_ws.column_dimensions["H"].width = 7

                # Define column style
                columnRetestCampaignNTStyle = styles.NamedStyle(
                    name="column_retestcampaignnt_style"
                )
                columnRetestCampaignNTStyle.alignment = styles.Alignment(
                    horizontal="center"
                )
                columnRetestCampaignNTStyle.font = styles.Font(
                    name="OCR A Extended", color="FFFFFF"
                )
                columnRetestCampaignNTStyle.fill = styles.PatternFill(
                    patternType="solid", fgColor="36A2EB"
                )
                columnRetestCampaignNAStyle = styles.NamedStyle(
                    name="column_retestcampaignna_style"
                )
                columnRetestCampaignNAStyle.alignment = styles.Alignment(
                    horizontal="center"
                )
                columnRetestCampaignNAStyle.font = styles.Font(
                    name="OCR A Extended", color="FFFFFF"
                )
                columnRetestCampaignNAStyle.fill = styles.PatternFill(
                    patternType="solid", fgColor="E0E0E0"
                )
                columnRetestCampaignNFStyle = styles.NamedStyle(
                    name="column_retestcampaignnf_style"
                )
                columnRetestCampaignNFStyle.alignment = styles.Alignment(
                    horizontal="center"
                )
                columnRetestCampaignNFStyle.font = styles.Font(
                    name="OCR A Extended", color="FFFFFF"
                )
                columnRetestCampaignNFStyle.fill = styles.PatternFill(
                    patternType="solid", fgColor="FF3333"
                )
                columnRetestCampaignPFStyle = styles.NamedStyle(
                    name="column_retestcampaignpf_style"
                )
                columnRetestCampaignPFStyle.alignment = styles.Alignment(
                    horizontal="center"
                )
                columnRetestCampaignPFStyle.font = styles.Font(
                    name="OCR A Extended", color="FFFFFF"
                )
                columnRetestCampaignPFStyle.fill = styles.PatternFill(
                    patternType="solid", fgColor="FFB266"
                )
                columnRetestCampaignFtyle = styles.NamedStyle(
                    name="column_retestcampaignf_style"
                )
                columnRetestCampaignFtyle.alignment = styles.Alignment(
                    horizontal="center"
                )
                columnRetestCampaignFtyle.font = styles.Font(
                    name="OCR A Extended", color="FFFFFF"
                )
                columnRetestCampaignFtyle.fill = styles.PatternFill(
                    patternType="solid", fgColor="33FF33"
                )

                # Add legend data.
                retestcampaigns_ws["A1"] = "Retest Campaign Summary"
                retestcampaigns_ws["A2"] = (
                    "NT: Not Tested, NA: Not Applicable, NF: Not Fixed, PF: Partially Fixed, F: Fixed"
                )

                # Beautify legend data
                retestcampaigns_ws.merge_cells("A1:H1")
                retestcampaigns_ws.merge_cells("A2:H2")

                retestcampaigns_ws["A1"].style = projectHeaderStyle
                retestcampaigns_ws["A2"].style = projectValueStyle

                # Add column headers.
                retestcampaigns_ws["A4"] = "Name"
                retestcampaigns_ws["B4"] = "Start Date"
                retestcampaigns_ws["C4"] = "End Date"
                retestcampaigns_ws["D4"] = "# NT"
                retestcampaigns_ws["E4"] = "# NA"
                retestcampaigns_ws["F4"] = "# NF"
                retestcampaigns_ws["G4"] = "# PF"
                retestcampaigns_ws["H4"] = "# F"

                # Beautify column headers.
                retestcampaigns_ws["A4"].style = columnHeaderStyle
                retestcampaigns_ws["B4"].style = columnHeaderStyle
                retestcampaigns_ws["C4"].style = columnHeaderStyle
                retestcampaigns_ws["D4"].style = columnHeaderStyle
                retestcampaigns_ws["E4"].style = columnHeaderStyle
                retestcampaigns_ws["F4"].style = columnHeaderStyle
                retestcampaigns_ws["G4"].style = columnHeaderStyle
                retestcampaigns_ws["H4"].style = columnHeaderStyle

                # Insert data as retest campaign summary.
                line = 5
                for retestcampaign in project.retestcampaign_set.all():
                    retestcampaigns_ws.cell(row=line, column=1).value = (
                        retestcampaign.name
                    )
                    retestcampaigns_ws.cell(row=line, column=2).value = (
                        retestcampaign.start_date
                    )
                    retestcampaigns_ws.cell(row=line, column=3).value = (
                        retestcampaign.end_date
                    )
                    retestcampaigns_ws.cell(row=line, column=4).value = len(
                        retestcampaign.get_not_tested_hits()
                    )
                    retestcampaigns_ws.cell(row=line, column=5).value = len(
                        retestcampaign.get_not_applicable_hits()
                    )
                    retestcampaigns_ws.cell(row=line, column=6).value = len(
                        retestcampaign.get_not_fixed_hits()
                    )
                    retestcampaigns_ws.cell(row=line, column=7).value = len(
                        retestcampaign.get_partially_fixed_hits()
                    )
                    retestcampaigns_ws.cell(row=line, column=8).value = len(
                        retestcampaign.get_fixed_hits()
                    )
                    retestcampaigns_ws.cell(row=line, column=4).style = (
                        columnRetestCampaignNTStyle
                    )
                    retestcampaigns_ws.cell(row=line, column=5).style = (
                        columnRetestCampaignNAStyle
                    )
                    retestcampaigns_ws.cell(row=line, column=6).style = (
                        columnRetestCampaignNFStyle
                    )
                    retestcampaigns_ws.cell(row=line, column=7).style = (
                        columnRetestCampaignPFStyle
                    )
                    retestcampaigns_ws.cell(row=line, column=8).style = (
                        columnRetestCampaignFtyle
                    )
                    line = line + 1

                # --------------------------------------------------------------------------
                # Create a specific sheet for each retest campaign.
                # --------------------------------------------------------------------------
                for retestcampaign in project.retestcampaign_set.all():
                    retestcampaign_ws = wb.create_sheet()
                    retestcampaign_ws.title = retestcampaign.name

                    # Define column size
                    retestcampaign_ws.column_dimensions["A"].width = 28
                    retestcampaign_ws.column_dimensions["B"].width = 10
                    retestcampaign_ws.column_dimensions["C"].width = 10
                    retestcampaign_ws.column_dimensions["D"].width = 17
                    retestcampaign_ws.column_dimensions["E"].width = 50
                    retestcampaign_ws.column_dimensions["F"].width = 30
                    retestcampaign_ws.column_dimensions["G"].width = 12
                    retestcampaign_ws.column_dimensions["H"].width = 17
                    retestcampaign_ws.column_dimensions["I"].width = 50
                    retestcampaign_ws.column_dimensions["J"].width = 70

                    # Add project data.
                    retestcampaign_ws["A1"] = "Campaign Name:"
                    retestcampaign_ws["A2"] = "Date:"
                    retestcampaign_ws["B1"] = retestcampaign.name
                    if (
                        retestcampaign.start_date is not None
                        and retestcampaign.end_date is not None
                    ):
                        retestcampaign_ws["B2"] = (
                            "From "
                            + str(retestcampaign.start_date)
                            + " To "
                            + str(retestcampaign.end_date)
                        )
                    else:
                        retestcampaign_ws["B2"] = "Not Defined"

                    # Beautify project data
                    retestcampaign_ws.merge_cells("B1:J1")
                    retestcampaign_ws.merge_cells("B2:J2")

                    retestcampaign_ws["A1"].style = projectHeaderStyle
                    retestcampaign_ws["A2"].style = projectHeaderStyle
                    retestcampaign_ws["B1"].style = projectValueStyle
                    retestcampaign_ws["B2"].style = projectValueStyle

                    # Add column headers.
                    retestcampaign_ws["A4"] = "Assessment"
                    retestcampaign_ws["B4"] = "Sev"
                    retestcampaign_ws["C4"] = (
                        "CVSSv3" if project.cvss_type == 3 else "CVSSv4"
                    )
                    retestcampaign_ws["D4"] = "ID"
                    retestcampaign_ws["E4"] = "Title"
                    retestcampaign_ws["F4"] = "Asset"
                    retestcampaign_ws["G4"] = "Fix Compl."
                    retestcampaign_ws["H4"] = "Status"
                    retestcampaign_ws["I4"] = "Labels"
                    retestcampaign_ws["J4"] = "CWE(s)"

                    # Beautify column headers.
                    retestcampaign_ws["A4"].style = columnHeaderStyle
                    retestcampaign_ws["B4"].style = columnHeaderStyle
                    retestcampaign_ws["C4"].style = columnHeaderStyle
                    retestcampaign_ws["D4"].style = columnHeaderStyle
                    retestcampaign_ws["E4"].style = columnHeaderStyle
                    retestcampaign_ws["F4"].style = columnHeaderStyle
                    retestcampaign_ws["G4"].style = columnHeaderStyle
                    retestcampaign_ws["H4"].style = columnHeaderStyle
                    retestcampaign_ws["I4"].style = columnHeaderStyle
                    retestcampaign_ws["J4"].style = columnHeaderStyle

                    line = 5
                    for retest_hit in retestcampaign.retesthit_set.all():
                        retestcampaign_ws.cell(row=line, column=1).value = (
                            retest_hit.hit.assessment.name
                        )
                        retestcampaign_ws.cell(row=line, column=2).value = "P{}".format(
                            retest_hit.hit.severity
                        )
                        retestcampaign_ws.cell(row=line, column=3).value = (
                            retest_hit.hit.get_cvss_value()
                        )
                        retestcampaign_ws.cell(row=line, column=4).value = (
                            retest_hit.hit.get_unique_id()
                        )
                        retestcampaign_ws.cell(row=line, column=5).value = (
                            retest_hit.hit.title
                        )
                        retestcampaign_ws.cell(row=line, column=6).value = (
                            retest_hit.hit.asset
                        )
                        retestcampaign_ws.cell(row=line, column=7).value = (
                            retest_hit.hit.get_fix_complexity_str()
                        )
                        retestcampaign_ws.cell(row=line, column=8).value = (
                            retest_hit.get_status_display()
                        )

                        label_str = ""
                        previous = ""
                        for label in retest_hit.hit.labels.all():
                            label_str = "{}{}{}".format(
                                label_str, previous, label.title
                            )
                            previous = "\n"
                        retestcampaign_ws.cell(row=line, column=9).value = label_str

                        cwe_str = ""
                        previous = ""
                        for cwe in retest_hit.hit.cwes.all():
                            cwe_str = "{}{}{}".format(cwe_str, previous, cwe)
                            previous = "\n"
                        retestcampaign_ws.cell(row=line, column=10).value = cwe_str

                        # Apply style from value.
                        if retest_hit.hit.severity == 1:
                            retestcampaign_ws.cell(row=line, column=2).style = (
                                criticalStyle
                            )
                        elif retest_hit.hit.severity == 2:
                            retestcampaign_ws.cell(row=line, column=2).style = highStyle
                        elif retest_hit.hit.severity == 3:
                            retestcampaign_ws.cell(row=line, column=2).style = (
                                mediumStyle
                            )
                        elif retest_hit.hit.severity == 4:
                            retestcampaign_ws.cell(row=line, column=2).style = lowStyle
                        elif retest_hit.hit.severity == 5:
                            retestcampaign_ws.cell(row=line, column=2).style = infoStyle

                        if retest_hit.hit.fix_complexity == 1:
                            retestcampaign_ws.cell(row=line, column=7).style = highStyle
                        elif retest_hit.hit.fix_complexity == 2:
                            retestcampaign_ws.cell(row=line, column=7).style = (
                                mediumStyle
                            )
                        elif retest_hit.hit.fix_complexity == 3:
                            retestcampaign_ws.cell(row=line, column=7).style = lowStyle
                        else:
                            retestcampaign_ws.cell(row=line, column=7).style = infoStyle

                        if retest_hit.status == "NT":
                            retestcampaign_ws.cell(row=line, column=8).style = (
                                columnRetestCampaignNTStyle
                            )
                        elif retest_hit.status == "NA":
                            retestcampaign_ws.cell(row=line, column=8).style = (
                                columnRetestCampaignNAStyle
                            )
                        elif retest_hit.status == "NF":
                            retestcampaign_ws.cell(row=line, column=8).style = (
                                columnRetestCampaignNFStyle
                            )
                        elif retest_hit.status == "PF":
                            retestcampaign_ws.cell(row=line, column=8).style = (
                                columnRetestCampaignPFStyle
                            )
                        else:
                            retestcampaign_ws.cell(row=line, column=8).style = (
                                columnRetestCampaignFtyle
                            )

                        # For the moment, CVSS3 and CVSS4 share the same classification values.
                        try:
                            if float(retest_hit.hit.get_cvss_value()) < 0.1:
                                retestcampaign_ws.cell(row=line, column=3).style = (
                                    infoStyle
                                )
                            elif float(retest_hit.hit.get_cvss_value()) < 4.0:
                                retestcampaign_ws.cell(row=line, column=3).style = (
                                    lowStyle
                                )
                            elif float(retest_hit.hit.get_cvss_value()) < 7.0:
                                retestcampaign_ws.cell(row=line, column=3).style = (
                                    mediumStyle
                                )
                            elif float(retest_hit.hit.get_cvss_value()) < 9.0:
                                retestcampaign_ws.cell(row=line, column=3).style = (
                                    highStyle
                                )
                            else:
                                retestcampaign_ws.cell(row=line, column=3).style = (
                                    criticalStyle
                                )
                        except ValueError:
                            retestcampaign_ws.cell(row=line, column=3).style = infoStyle

                        retestcampaign_ws.cell(row=line, column=1).alignment = (
                            Alignment(vertical="center", horizontal="left")
                        )
                        retestcampaign_ws.cell(row=line, column=2).alignment = (
                            Alignment(vertical="center", horizontal="center")
                        )
                        retestcampaign_ws.cell(row=line, column=3).alignment = (
                            Alignment(vertical="center", horizontal="center")
                        )
                        retestcampaign_ws.cell(row=line, column=4).alignment = (
                            Alignment(vertical="center", horizontal="center")
                        )
                        retestcampaign_ws.cell(row=line, column=5).alignment = (
                            Alignment(vertical="top", horizontal="left", wrap_text=True)
                        )
                        retestcampaign_ws.cell(row=line, column=6).alignment = (
                            Alignment(vertical="top", horizontal="left", wrap_text=True)
                        )
                        retestcampaign_ws.cell(row=line, column=7).alignment = (
                            Alignment(vertical="center", horizontal="center")
                        )
                        retestcampaign_ws.cell(row=line, column=8).alignment = (
                            Alignment(vertical="center", horizontal="center")
                        )
                        retestcampaign_ws.cell(row=line, column=9).alignment = (
                            Alignment(vertical="top", wrap_text=True)
                        )
                        retestcampaign_ws.cell(row=line, column=10).alignment = (
                            Alignment(vertical="top", wrap_text=True)
                        )

                        line = line + 1

                    retestcampaign_ws.auto_filter.ref = "A4:H{}".format(line)

            # Prepare HTTP response.
            response = HttpResponse(save_virtual_workbook(wb),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = (
                "attachment; filename=" + project.name + ".xlsx"
            )
            wb.close()
        else:
            response = Response(status=status.HTTP_403_FORBIDDEN)
    except Flag.DoesNotExist:
        response = Response(status=status.HTTP_404_NOT_FOUND)
    return response


@csrf_exempt
@ptart_authentication
@api_view(http_method_names=["GET"])
@renderer_classes([ZipRenderer])
def project_latex(request, pk):
    response = None
    try:
        project = Project.objects.get(pk=pk)
        if project.is_user_can_view(request.user):
            response = HttpResponse()
            zf = zipfile.ZipFile(response, "w")

            # Retrieve Screenshots.
            for assessment in project.assessment_set.all():
                for hit in assessment.displayable_hits():
                    for screenshot in hit.screenshot_set.all():
                        ##Screenshots are stored in the zip file.
                        raw_data = screenshot.get_raw_data()
                        if raw_data is not None:
                            zf.writestr(
                                "screenshots/{}.png".format(screenshot.id), raw_data
                            )
                        else:
                            with open(
                                "reports/resources/default_image.png", "rb"
                            ) as file:
                                default_image_data = file.read()
                                zf.writestr(
                                    "screenshots/{}.png".format(screenshot.id),
                                    default_image_data,
                                )

            # Retrieve Screenshots for retest campaignes.
            for retestcampaign in project.retestcampaign_set.all():
                for retest_hit in retestcampaign.retesthit_set.all():
                    for screenshot in retest_hit.retestscreenshot_set.all():
                        ##Screenshots are stored in the zip file.
                        raw_data = screenshot.get_raw_data()
                        if raw_data is not None:
                            zf.writestr(
                                "screenshots_retest/{}.png".format(screenshot.id),
                                raw_data,
                            )
                        else:
                            with open(
                                "reports/resources/default_image.png", "rb"
                            ) as file:
                                default_image_data = file.read()
                                zf.writestr(
                                    "screenshots_retest/{}.png".format(screenshot.id),
                                    default_image_data,
                                )

            # Retrieve AttackScenarios.
            for attackscenario in project.attackscenario_set.all():
                # SVG must be converted to PNG to facilitate integration in LaTeX report.
                tempSvgFilename = "{}_{}_{}.svg".format(
                    project.id, attackscenario.id, random.randint(0, 10000)
                )
                tempPngFilename = "{}_{}_{}.png".format(
                    project.id, attackscenario.id, random.randint(0, 10000)
                )
                
                with open(tempSvgFilename,"w") as tempSvgFile:
                    tempSvgFile.write(attackscenario.svg)
                    tempSvgFile.close()
                    
                pngContent = cairosvg.svg2png(
                    url=tempSvgFilename
                )
                
                with open(tempPngFilename,"wb") as tempPngFile:
                    tempPngFile.write(pngContent)
                    tempPngFile.close()
                    
                zf.write(
                    tempPngFilename, "attackscenarios/{}.png".format(attackscenario.id)
                )
                
                os.remove(tempSvgFilename)
                os.remove(tempPngFilename)

            # Add resources for LaTeX
            zf.write("reports/resources/companylogo.png", "resources/companylogo.png")
            zf.write("reports/resources/logo.png", "resources/logo.png")

            # Generate Latex report.
            # Custom environment is used to avoid syntax conflict between Jinja & LaTex
            with open("reports/report_latex.tex") as file_:
                env = jinja2.Environment(
                    block_start_string="\\BLOCK{",
                    block_end_string="}",
                    variable_start_string="\\VAR{",
                    variable_end_string="}",
                    comment_start_string="\\#{",
                    comment_end_string="}",
                    line_statement_prefix="%%",
                    line_comment_prefix="%#",
                    trim_blocks=True,
                    autoescape=False,
                    loader=jinja2.FileSystemLoader(os.path.abspath(".")),
                )

                # Filters that has been used in the LaTex Report.
                escape_tex_table = {
                    "&": r"\&",
                    "%": r"\%",
                    "$": r"\$",
                    "#": r"\#",
                    "_": r"\_",
                    "~": r"\textasciitilde{}",
                    "^": r"\^{}",
                    "\\": r"\textbackslash{}",
                }
                tex_regex = re.compile(
                    "|".join(
                        re.escape(str(key))
                        for key in sorted(
                            escape_tex_table.keys(), key=lambda item: -len(item)
                        )
                    )
                )
                screenshot_pattern = re.compile(r"(\\pandocbounded\{.*?\}\})")

                def tex_escape(text):
                    return tex_regex.sub(
                        lambda match: escape_tex_table[match.group()], text
                    )

                def markdown_to_latex_common(md, folder, screenshot_class):
                    latex = pypandoc.convert_text(
                        md,
                        "latex",
                        format="md",
                        extra_args=["--wrap=preserve", "--highlight-style=tango"],
                    )
                    if folder is not None:
                        matches = screenshot_pattern.findall(latex)
                        if matches is not None and len(matches) > 0:
                            # Clean the markdown from the screenshots.
                            latex = latex.replace("\n\\begin{figure}\n\\centering", "")
                            latex = latex.replace(
                                "\\caption{ptart\\_screenshot}\n\\end{figure}\n", ""
                            )
                            for match in matches:
                                try:
                                    item = screenshot_class.objects.get(
                                        pk=match.split("/")[-2]
                                    )
                                    if item.is_user_can_view(request.user):
                                        graphic = """
                                        \\begin{{figure}}[H]
                                        \\centering
                                        \\includegraphics[max width=\\textwidth]{{./{}/{}.png}}
                                        \\caption{{{}}}
                                        \\end{{figure}}
                                        """.format(
                                            folder, item.id, tex_escape(item.caption)
                                        )
                                        latex = latex.replace(match, graphic)
                                    else:
                                        latex = latex.replace(match, "")
                                except Screenshot.DoesNotExist as e:
                                    latex = latex.replace(match, "")
                    return latex

                def markdown_to_latex(md):
                    return markdown_to_latex_common(md, None, None)

                def markdown_to_latex_hit(md):
                    return markdown_to_latex_common(md, "screenshots", Screenshot)

                def markdown_to_latex_retesthit(md):
                    return markdown_to_latex_common(
                        md, "screenshots_retest", RetestScreenshot
                    )

                # End of filters.
                env.filters["mdtolatex"] = markdown_to_latex
                env.filters["mdtolatexhit"] = markdown_to_latex_hit
                env.filters["mdtolatexretesthit"] = markdown_to_latex_retesthit
                env.filters["escape"] = tex_escape
                template = env.from_string(file_.read())
                zf.writestr(
                    "{}.tex".format(project.name).replace(" ", "_"),
                    template.render(
                        project=project, labels=Label.get_viewable(request.user)
                    ),
                )

            # Prepare HTTP response.
            response.content_type = "application/zip"
            response["Content-Disposition"] = (
                "attachment; filename=" + project.name + ".zip"
            )
        else:
            response = Response(status=status.HTTP_403_FORBIDDEN)
    except Flag.DoesNotExist:
        response = Response(status=status.HTTP_404_NOT_FOUND)
    return response


@csrf_exempt
@ptart_authentication
@api_view(http_method_names=["GET"])
def project_json(request, pk):
    response = None
    try:
        project = Project.objects.get(pk=pk)
        if project.is_user_can_view(request.user):
            data = {
                "name": project.name,
                "executive_summary": project.executive_summary,
                "engagement_overview": project.engagement_overview,
                "conclusion": project.conclusion,
                "scope": project.scope,
                "client": project.client,
                "start_date": project.start_date,
                "end_date": project.end_date,
                "cvss_type": project.cvss_type,
                "cwe_version": project.cwes.version,
                "tools": [tool.name for tool in project.tools.all()],
                "methodologies": [
                    methodology.name for methodology in project.methodologies.all()
                ],
                "pentesters": [
                    {
                        "username": pentester.username,
                        "first_name": pentester.first_name,
                        "last_name": pentester.last_name,
                    }
                    for pentester in project.pentesters.all()
                ],
                "reviewers": [
                    {
                        "username": reviewer.username,
                        "first_name": reviewer.first_name,
                        "last_name": reviewer.last_name,
                    }
                    for reviewer in project.reviewers.all()
                ],
                "assessments": [
                    {
                        "title": assessment.name,
                        "hits": [
                            {
                                "id": hit.get_unique_id(),
                                "title": hit.title,
                                "body": hit.body,
                                "remediation": hit.remediation,
                                "asset": hit.asset,
                                "severity": hit.severity,
                                "fix_complexity": hit.fix_complexity,
                                "cvss_vector": (
                                    (
                                        hit.cvss3.get_cvss_string()
                                        if hit.cvss3 is not None
                                        else ""
                                    )
                                    if project.cvss_type == 3
                                    else (
                                        hit.cvss4.get_cvss_string()
                                        if hit.cvss4 is not None
                                        else ""
                                    )
                                ),
                                "cvss_score": (
                                    (
                                        hit.cvss3.decimal_value
                                        if hit.cvss3 is not None
                                        else ""
                                    )
                                    if project.cvss_type == 3
                                    else (
                                        hit.cvss4.decimal_value
                                        if hit.cvss4 is not None
                                        else ""
                                    )
                                ),
                                "added": hit.added,
                                "labels": [
                                    label_hit.title for label_hit in hit.labels.all()
                                ],
                                "cwes": [
                                    cwe_hit.export() for cwe_hit in hit.cwes.all()
                                ],
                                "screenshots": [
                                    screenshot.export()
                                    for screenshot in hit.screenshot_set.all()
                                ],
                                "attachments": [
                                    attachment.export()
                                    for attachment in hit.attachment_set.all()
                                ],
                                "references": [
                                    reference.export()
                                    for reference in hit.hitreference_set.all()
                                ],
                            }
                            for hit in assessment.displayable_hits()
                        ],
                    }
                    for assessment in project.assessment_set.all()
                ],
                "retests": [
                    {
                        "name": retestcampaign.name,
                        "introduction": retestcampaign.introduction,
                        "conclusion": retestcampaign.conclusion,
                        "start_date": retestcampaign.start_date,
                        "end_date": retestcampaign.end_date,
                        "hits": [
                            {
                                "id": f"{retesthit.hit.get_unique_id()}-RT",
                                "status": retesthit.status,
                                "body": retesthit.body,
                                "original_hit": {
                                    "id": retesthit.hit.get_unique_id(),
                                    "title": retesthit.hit.title,
                                    "body": retesthit.hit.body,
                                    "remediation": retesthit.hit.remediation,
                                    "asset": retesthit.hit.asset,
                                    "severity": retesthit.hit.severity,
                                    "fix_complexity": retesthit.hit.fix_complexity,
                                    "cvss_vector": (
                                        (
                                            retesthit.hit.cvss3.get_cvss_string()
                                            if retesthit.hit.cvss3 is not None
                                            else ""
                                        )
                                        if project.cvss_type == 3
                                        else (
                                            retesthit.hit.cvss4.get_cvss_string()
                                            if retesthit.hit.cvss4 is not None
                                            else ""
                                        )
                                    ),
                                    "cvss_score": (
                                        (
                                            retesthit.hit.cvss3.decimal_value
                                            if retesthit.hit.cvss3 is not None
                                            else ""
                                        )
                                        if project.cvss_type == 3
                                        else (
                                            retesthit.hit.cvss4.decimal_value
                                            if retesthit.hit.cvss4 is not None
                                            else ""
                                        )
                                    ),
                                    "added": retesthit.hit.added,
                                    "labels": [
                                        label_hit.title
                                        for label_hit in retesthit.hit.labels.all()
                                    ],
                                    "cwes": [
                                        cwe_hit.export()
                                        for cwe_hit in retesthit.hit.cwes.all()
                                    ],
                                },
                                "screenshots": [
                                    screenshot.export()
                                    for screenshot in retesthit.retestscreenshot_set.all()
                                ],
                            }
                            for retesthit in retestcampaign.retesthit_set.all()
                        ],
                    }
                    for retestcampaign in project.retestcampaign_set.all()
                ],
            }

            # Prepare HTTP response.
            response = JsonResponse(data=data)
            response.content_type = "application/json"
            response["Content-Disposition"] = (
                "attachment; filename=" + project.name + ".json"
            )
            response.accepted_media_type = "application/json"
            response.renderer_context = {}
        else:
            response = Response(status=status.HTTP_403_FORBIDDEN)
    except Flag.DoesNotExist:
        response = Response(status=status.HTTP_404_NOT_FOUND)
    return response


@csrf_exempt
@ptart_authentication
@api_view(http_method_names=["GET"])
@renderer_classes([TemplateHTMLRenderer])
def project_slides(request, pk):
    """
    Generate a reveal.js HTML presentation for a project.
    Returns a standalone HTML file that can be opened in a browser.
    Add ?print-pdf to the URL when opening the file to enable PDF generation mode.
    """
    response = None
    try:
        project = Project.objects.get(pk=pk)
        if project.is_user_can_view(request.user):
            from django.template.loader import render_to_string

            # Render the reveal.js template with the project context
            html_content = render_to_string(
                "projects/project-report.html",
                {"project": project}
            )

            # Prepare HTTP response with HTML content
            response = HttpResponse(html_content, content_type="text/html")
        else:
            response = Response(status=status.HTTP_403_FORBIDDEN)
    except Project.DoesNotExist:
        response = Response(status=status.HTTP_404_NOT_FOUND)
    return response


@csrf_exempt
@ptart_authentication
@api_view(["POST"])
def chatgpt(request):
    response = None

    if settings.CHATGPT_API_KEY != "NotConfigured":
        if request.data["question"] != None:

            req_response = requests.post(
                "https://api.openai.com/v1/completions",
                json={
                    "prompt": request.data["question"],
                    "max_tokens": 2048,
                    "model": "gpt-3.5-turbo-instruct",
                },
                headers={
                    "Authorization": "Bearer " + settings.CHATGPT_API_KEY,
                    "Content-Type": "application/json",
                },
            )
            if req_response.status_code == 200:
                response = Response(
                    req_response.json()["choices"][0]["text"], status=status.HTTP_200_OK
                )
            else:
                response = Response({}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        else:
            response = Response({}, status=status.HTTP_400_BAD_REQUEST)
    else:
        response = Response({}, status=status.HTTP_501_NOT_IMPLEMENTED)

    return response


@csrf_exempt
@ptart_authentication
@action(methods=["GET"], detail=True)
def audit(request):
    response = None
    if not request.user.is_staff:
        response = Response(status=status.HTTP_403_FORBIDDEN)
    else:
        try:

            wb = Workbook()
            ws = wb.active
            ws.title = "Audit"

            # Define column size
            wb.active.column_dimensions["A"].width = 28
            wb.active.column_dimensions["B"].width = 15
            wb.active.column_dimensions["C"].width = 14
            wb.active.column_dimensions["D"].width = 14
            wb.active.column_dimensions["E"].width = 50
            wb.active.column_dimensions["F"].width = 50
            wb.active.column_dimensions["G"].width = 9
            wb.active.column_dimensions["H"].width = 9
            wb.active.column_dimensions["I"].width = 9
            wb.active.column_dimensions["J"].width = 9
            wb.active.column_dimensions["K"].width = 9

            # Add project data.
            ws["A1"] = "Audit Report"
            ws["A2"] = "Date:"
            ws["A3"] = "User:"
            ws["B2"] = datetime.datetime.now()
            ws["B3"] = "{} - {} {}".format(
                request.user.username, request.user.first_name, request.user.last_name
            )

            # Beautify project data
            ws.merge_cells("B1:K1")
            ws.merge_cells("B2:K2")
            ws.merge_cells("B3:K3")

            projectHeaderStyle = styles.NamedStyle(name="project_header_style")
            projectHeaderStyle.font = styles.Font(
                name="Calibri", size=14, bold=True, color="000000"
            )
            projectHeaderStyle.fill = styles.PatternFill(
                patternType="solid", fgColor="00B0F0"
            )
            projectHeaderStyle.alignment = styles.Alignment(horizontal="left")
            ws["A1"].style = projectHeaderStyle
            ws["A2"].style = projectHeaderStyle
            ws["A3"].style = projectHeaderStyle

            projectValueStyle = styles.NamedStyle(name="project_value_style")
            projectValueStyle.font = styles.Font(
                name="Calibri", size=14, italic=True, color="000000"
            )
            projectValueStyle.fill = styles.PatternFill(
                patternType="solid", fgColor="00B0F0"
            )
            projectValueStyle.alignment = styles.Alignment(horizontal="left")

            ws["B1"].style = projectValueStyle
            ws["B2"].style = projectValueStyle
            ws["B3"].style = projectValueStyle
            ws["B2"].number_format = "YYYY MMM DD"

            # Add column header.
            ws["A5"] = "Project"
            ws["B5"] = "Client"
            ws["C5"] = "Start Date"
            ws["D5"] = "End Date"
            ws["E5"] = "Pentesters"
            ws["F5"] = "Reviewers"
            ws["G5"] = "#P1"
            ws["H5"] = "#P2"
            ws["I5"] = "#P3"
            ws["J5"] = "#P4"
            ws["K5"] = "#P5"

            columnHeaderStyle = styles.NamedStyle(name="column_header_style")
            columnHeaderStyle.font = styles.Font(
                name="Calibri", size=12, bold=True, color="000000"
            )
            columnHeaderStyle.fill = styles.PatternFill(
                patternType="solid", fgColor="92D050"
            )
            columnHeaderStyle.alignment = styles.Alignment(horizontal="center")

            ws["A5"].style = columnHeaderStyle
            ws["B5"].style = columnHeaderStyle
            ws["C5"].style = columnHeaderStyle
            ws["D5"].style = columnHeaderStyle
            ws["E5"].style = columnHeaderStyle
            ws["F5"].style = columnHeaderStyle
            ws["G5"].style = columnHeaderStyle
            ws["H5"].style = columnHeaderStyle
            ws["I5"].style = columnHeaderStyle
            ws["J5"].style = columnHeaderStyle
            ws["K5"].style = columnHeaderStyle

            line = 6

            for project in Project.objects.all().order_by("added"):

                ws.cell(row=line, column=1).value = project.name
                ws.cell(row=line, column=2).value = project.client
                ws["C{}".format(line)].number_format = "YYYY MMM DD"
                ws["D{}".format(line)].number_format = "YYYY MMM DD"
                ws.cell(row=line, column=3).value = project.start_date
                ws.cell(row=line, column=4).value = project.end_date

                # Construct the auditor string
                pentester_str = ""
                previous = ""
                for pentester in project.pentesters.all():
                    pentester_str = "{}{}{}".format(
                        pentester_str, previous, pentester.username
                    )
                    previous = ", "
                ws.cell(row=line, column=5).value = pentester_str

                # Construct the reviewer string
                reviewer_str = ""
                previous = ""
                for reviewer in project.reviewers.all():
                    reviewer_str = "{}{}{}".format(
                        reviewer_str, previous, reviewer.username
                    )
                    previous = ", "
                ws.cell(row=line, column=6).value = reviewer_str

                ws.cell(row=line, column=7).value = len(project.p1_hits())
                ws.cell(row=line, column=8).value = len(project.p2_hits())
                ws.cell(row=line, column=9).value = len(project.p3_hits())
                ws.cell(row=line, column=10).value = len(project.p4_hits())
                ws.cell(row=line, column=11).value = len(project.p5_hits())

                line = line + 1

            ws.auto_filter.ref = "A5:K{}".format(line)

            # Prepare HTTP response.
            response = Response(save_virtual_workbook(wb))
            response.content_type = (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = 'attachment; filename="audit.xlsx"'
            response.accepted_renderer = BinaryRenderer()
            response.accepted_media_type = (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response.renderer_context = {}

            wb.close()
        except Flag.DoesNotExist:
            response = Response(status=status.HTTP_404_NOT_FOUND)
    return response


@csrf_exempt
@ptart_authentication
@api_view(["POST", "DELETE"])
def manage_token(request):
    """
    Grant or Revoke authentication token.
    """
    response = Response(status=status.HTTP_404_NOT_FOUND)
    if request.method == "POST":
        response = Response(status=status.HTTP_404_NOT_FOUND)
        if not Token.objects.filter(user=request.user).exists():
            token = Token.objects.create(user=request.user)
            response = Response({"token": token.key}, status=status.HTTP_201_CREATED)
    else:
        response = Response(status=status.HTTP_404_NOT_FOUND)
        if Token.objects.filter(user=request.user).exists():
            Token.objects.filter(user=request.user).delete()
            response = Response({"token": ""}, status=status.HTTP_204_NO_CONTENT)

    return response


@csrf_exempt
@ptart_authentication
@api_view(["POST"])
def change_password(request):
    """
    Change current user password.
    """
    response = Response(
        {"The current password is not valid"}, status=status.HTTP_400_BAD_REQUEST
    )
    user = authenticate(
        request, username=request.user, password=request.data["oldPassword"]
    )
    if user is not None:
        password1 = str(request.data["newPassword1"])
        password2 = str(request.data["newPassword2"])
        if password1 == password2:
            try:
                validate_password(password1, user)
                user.set_password(password1)
                user.save()
                response = Response({}, status=status.HTTP_204_NO_CONTENT)
            except ValidationError as err:
                response = Response(err, status=status.HTTP_400_BAD_REQUEST)
        else:
            response = Response(
                {"New passwords are not matching"}, status=status.HTTP_400_BAD_REQUEST
            )
    return response


#
# CRUD operations for a specific item.
#
def item(request, pk, class_name, serializer_name):
    response = None
    try:
        item = class_name.objects.get(pk=pk)

        if request.method == "GET":
            if item.is_user_can_view(request.user):
                response = Response(serializer_name(item).data)
            else:
                response = Response(status=status.HTTP_403_FORBIDDEN)
        elif request.method == "PUT" or request.method == "PATCH":
            serializer = serializer_name(
                item, data=request.data, partial=(request.method == "PATCH")
            )
            if serializer.is_valid():
                if item.is_user_can_edit(request.user):
                    serializer.save()
                    response = Response(serializer.data)
                else:
                    response = Response(status=status.HTTP_403_FORBIDDEN)
            else:
                response = Response(
                    serializer.errors, status=status.HTTP_400_BAD_REQUEST
                )

        elif request.method == "DELETE":
            if item.is_user_can_edit(request.user):
                response = Response(
                    serializer_name(item).data, status=status.HTTP_200_OK
                )
                item.delete()
            else:
                response = Response(status=status.HTTP_403_FORBIDDEN)

    except class_name.DoesNotExist:
        response = Response(status=status.HTTP_404_NOT_FOUND)

    return response


#
# CRUD operations for a specific list of items.
#
def items(request, class_name, serializer_name):
    response = None
    if request.method == "GET":
        response = Response(
            serializer_name(class_name.get_viewable(request.user), many=True).data
        )
    elif request.method == "POST":
        serializer = serializer_name(data=request.data)
        if serializer.is_valid():
            if class_name(**serializer.validated_data).is_user_can_create(request.user):
                serializer.save()
                response = Response(serializer.data, status=status.HTTP_201_CREATED)
            else:
                response = Response(status=status.HTTP_403_FORBIDDEN)
        else:
            response = Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    return response